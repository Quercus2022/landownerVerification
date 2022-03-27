#! python3
# testScraper.py - Parse HTML for property report webpage for deeded owner and their address.
import openpyxl
from openpyxl.styles import PatternFill
import requests
from bs4 import BeautifulSoup
import re


# gets a list of last names from database Excel
def getLastNameList(wsName):
    x = 1
    cells = []
    for i in wsName:
        cell = wsName.cell(row=x, column=6).value
        cells.append(cell)
        x += 1
        if i is False: print(i)
    return cells


# get list of street address from database Excel
def getAddressLst(wsAdr):
    x = 1
    cells = []
    for i in wsAdr:
        cell = wsAdr.cell(row=x, column=4).value
        cells.append(cell)
        x += 1
        if i is False: print(i)
    return cells


# get list of parcel IDs for database Excel
def getPID(ws):
    x = 1
    cells = []
    for i in ws:
        cell = ws.cell(row=x, column=22).value
        cells.append(cell)
        x += 1
        if i is False: print(i)
    return cells


# extracts for webpage a string with deeded landowners full names including trusts if present
def extractOwner(stew):
    # id is different depending on whether landowner name is a
    # link or not. Only  char difference so a regular expression
    # will take care of it.
    x = stew.find(id=re.compile(r"ctlBodyPane_ctl01_ctl01_lstOwner_ctl01_lnkOwnerName_l..Search"))
    if x is None:
        ownerA = 'none'
    else:
        ownerA = x.text  # get any text out of html file
    return ownerA


# extracts from webpage a string with deeded landowners mailing address
def extractAddress(stew):
    tdLst = stew.find(id='ctlBodyPane_ctl01_ctl01_lstOwner_ctl01_lblOwnerAddress')
    # <br> is within address txt
    if tdLst is None:
        ad4 = 'none'
    else:
        myTag = tdLst.find_all('br')
        # replace br with space
        myTag[1].replace_with(' ')
        # extract only the text from selected html
        ad1 = tdLst.text
        # split the text into address1(street address) and address2(city)
        ad2 = re.split(r'\w*,\s', ad1)
        ad3 = ad2[0].strip()  # selects street address and strips whitespace
        # removes all spaces from string for easier comparison
        ad4 = ad3.replace(' ', '')
    return ad4


# checks if last name from website matches the one in Excel, returns a Boolean
def lastNameCheck(a, b, c):
    name1 = b[c].split()
    # if a == name1[0]:  # Exact last name match.
    if a in name1:  # Is the database last name found anywhere in the deeded owners name.
        return True
    else:
        return False


# checks if address from webpage matches the one in Excel, returns a Boolean
def addressCheck(a, b, c):
    # remove all spaces from string for easier comparison
    adLst = b[c].replace(' ', '')
    # if a == adLst:  # Exact address match
    if a in adLst:  # Is the database address found anywhere in the website address.
        return True
    else:
        return False


# Color rows based on if the information is correct or not
def checkOutput(own, add, ws, i):
    x = i + 1
    if own is False:  # lastNameCheck returns true if the names match and false if they don't.
        # Fill row with red
        for cell in ws[x:x]:
            cell.fill = PatternFill(fill_type='solid', start_color='00FF0000', end_color='00FF0000')
    elif add is False:  # the addressCheck returns true if the address matches and false if it doesn't.
        # Fill row with red
        for cell in ws[x:x]:
            cell.fill = PatternFill(fill_type='solid', start_color='00FF0000', end_color='00FF0000')
    else:  # if lastNameCheck and addressCheck return True then the information is correct.
        # Fill row with green
        for cell in ws[x:x]:
            cell.fill = PatternFill(fill_type='solid', start_color='00008000', end_color='00008000')


########################################################################################


xlFile = input('Enter .xlsx file to open: ')  # reference workbook
try:
    wb1 = openpyxl.load_workbook(xlFile)  # open workbook
    ws1 = wb1['Sheet1']  # active sheet of database workbook
except:
    print('Error: ' + xlFile + ' failed to open')
    quit()

databaseLastNames = getLastNameList(ws1)  # get list of last names from Excel document
addressLst1 = getAddressLst(ws1)  # get list of address in column 1 (street address)
pID = getPID(ws1)  # get list of parcel IDs

# main loop
for p in range(1, ws1.max_row):
    url = 'https://beacon.schneidercorp.com/Application.aspx?AppID=129&' \
          'LayerID=1554&PageTypeID=4&PageID=817&KeyValue=' + str(pID[p])

    page = requests.get(url)
    soup = BeautifulSoup(page.content, 'html.parser')

    test = soup.find('p')
    if test is not None:
        print('Beacon has shut down scraper. Change IP.')
        quit()

    owner = extractOwner(soup).split()[0]  # extracts deeded owners name from HTML, splits the string into component
    # words, and assigns the one at index[0], their last name, to 'owner'.

    # extracts deeded owners mailing address from HTML, and splits at the end of street address.
    address = extractAddress(soup)

    lnCheck = lastNameCheck(owner, databaseLastNames, p)
    adCheck = addressCheck(address, addressLst1, p)

    checkOutput(lnCheck, adCheck, ws1, p)

    wb1.save('johnsonX.xlsx')

wb1.save('johnsonX.xlsx')
print('Done')
